import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.worksheet import Worksheet

from performance import SHEET_NAME


def plt_export_chart(excel_file, output):
    excel = load_workbook(excel_file)
    table = excel[SHEET_NAME]
    assert isinstance(table, Worksheet)
    rows = table.max_row

    datax = []
    datay = []
    i = 1
    for row in range(2, rows + 1):
        datax.append(table.cell(row, 1).value)
        datay.append(table.cell(row, 2).value)
        i += 1

    # 设置输出的图片大小
    figsize = 14, 9
    figure, ax = plt.subplots(figsize=figsize)

    plt.title("Total Memory Summary")
    plt.ylabel("Memory(KB)")
    plt.xlabel("Time")
    plt.plot(datax, datay, color='Red')
    plt.xticks(datax, __get_ticks(datax), rotation='vertical')
    plt.tick_params(labelsize=14)
    labels = ax.get_xticklabels() + ax.get_yticklabels()
    [label.set_fontname('Times New Roman') for label in labels]
    plt.savefig(output)
    plt.close('all')


def plt_export_history(pkg: str, histories: list, output: str):
    data_avg = []
    data_max = []
    datax = []

    for history in histories:
        if history.pkg != pkg:
            continue
        if not history.success:
            continue
        data_avg.append(history.avg)
        data_max.append(history.max)
        datax.append(history.date)

    # 设置输出的图片大小
    figsize = 14, 9
    figure, ax = plt.subplots(figsize=figsize)

    plt.title("History Memory Summary")
    plt.ylabel("Memory(KB)")
    plt.xlabel("Date")
    l1, = plt.plot(datax, data_max, color='Red')
    l2, = plt.plot(datax, data_avg, color='Blue')
    plt.xticks(datax, datax, rotation='vertical')
    plt.legend(handles=[l1, l2], labels=['max', 'avg'], loc='upper center', frameon=False)
    plt.tick_params(labelsize=12)
    labels = ax.get_xticklabels() + ax.get_yticklabels()
    [label.set_fontname('Times New Roman') for label in labels]
    plt.savefig(output)
    plt.close('all')


def __get_ticks(xs: list):
    show_len = 30
    count = len(xs)
    # 30取1
    if count < show_len:
        return xs
    length = int(count / show_len + 1)
    data = [xs[0]]
    for i in range(1, len(xs) - 1):
        if i % length == 0:
            data.append(xs[i])
        else:
            data.append('')
    data.append(xs[-1])
    return data

