#! /Users/xiaotian/anaconda3/bin/python3.6
#coding=utf-8

import tinydb as td
from datetime import datetime, timedelta
import sys
import re
import pyperclip
import os.path as ospath
import os
import json
from ComplexTDL import sysXXt
from ComplexTDL.CurrentPath import *

# init tinydb

db = td.TinyDB(BASEPATH + 'db.json')
taskDB = db.table('task')
relationshipDB = db.table('relationship')
logDB = db.table('log')

# read the job information

with open( BASEPATH + "WorkType.txt") as f:
    exec(f.read())


def GetPeriod(period = 'this week'):

    if period == 'this week':

        today = datetime.today()
        weekday = today.weekday()
        startDate = today - timedelta(days = weekday)
        startDateString = datetime.strftime(startDate, '%Y-%m-%d')
        endDate = today + timedelta(days = (6 - weekday))
        endDateString = datetime.strftime(endDate, '%Y-%m-%d')

    elif period == 'last week':

        today = datetime.today()
        weekday = today.weekday()
        startDate = today - timedelta(days = (weekday + 7))
        startDateString = datetime.strftime(startDate, '%Y-%m-%d')
        endDate = today - timedelta(days = (weekday + 1 ))
        endDateString = datetime.strftime(endDate, '%Y-%m-%d')

    elif period == 'next week':

        today = datetime.today()
        weekday = today.weekday()
        startDate = today + timedelta(days = (7 - weekday))
        startDateString = datetime.strftime(startDate, '%Y-%m-%d')
        endDate = today + timedelta(days = (13 - weekday))
        endDateString = datetime.strftime(endDate, '%Y-%m-%d')

    elif period == 'this month':

        thisMonth = '%02d'%datetime.today().month
        thisYear = str(datetime.today().year)
        startDateString = endDateString = thisYear + '-' + thisMonth

    elif period == 'last month':

        thisMonth = datetime.today().month
        thisYear = datetime.today().year
        if thisMonth != 1:
            theMonth = '%02d'%(thisMonth - 1)
            theYear = str(thisYear)
        else:
            theMonth = 12
            theYear = str(thisYear - 1)
        startDateString = endDateString = theYear + '-' + theMonth

    elif period == 'all':

        startDateString = '0'
        endDateString = '3'

    elif period == '':

        startDateString = ''
        endDateString = ''

    else: # an explict date

        t = period.split('|')
        try:
            if len(t) == 2:
                startDateString = t[0]
                endDateString = t[1]
            else:
                if re.search(r'[0-9]{4}-[0-9]{2}-[0-9]{2}', t[0]):
                    startDateString = endDateString = t[0]
                else:
                    delta = int(re.findall('[-]*[0-9].*', t[0])[0])
                    startDateString = endDateString = datetime.strftime(datetime.today() + timedelta(days = delta), '%Y-%m-%d')
        except:
            print('Parameters error, type -h for help')
            sys.exit()

    return startDateString, endDateString


def FindParent(parentNumber, list):
    for i in list:
        if i['no'] == parentNumber: return i
        elif 'ch' in i.keys():
            t = FindParent(parentNumber, i['ch'])
            if t != None: return t
    
    return None


# process description
def ProcessDescription(json, startDate = None, endDate = None):

    description = json['de']

    if startDate == None or endDate == None:
        result = description # des[0]
    else:

        dates = re.findall(r'[\n]*=== ([0-9]{4}-[0-9]{2}-[0-9]{2}) ===\n', description)
        description = re.sub(r'([\n]*=== [0-9]{4}-[0-9]{2}-[0-9]{2} ===\n)', '#@%@#', description)
        des = description.split('#@%@#')
        result = []
        
        i = 0
        for i in range(0, len(dates)):
            date = dates[i]
            if date >= startDate and date <= endDate:
                result.append( '=== ' + date + ' ===\n' + des[i+1])
        
        result = '---jnv8h line---'.join(result)

    if result == '': result = 'RT'

    return result


# Generate the report data
def Report(period = 'this week', range = 'todo', person = 'me', deadline = False, conditions = None, detail = 'some', sort = 'module'):
    # period: the time period you want to know, the period can be like:
    #     * last week / this week / next week
    #     * last month/ this month / next month
    #     * all
    #     * xxxx-xx-xx or xxxx-xx-xx|xxxx-xx-xx
    #     * n, where n= ..., -2, -1, 0, 1, 2,... 0 means today, -1 means yesterday and 1 means tomorrow
    # range: 
    #     * 'todo': things to do
    #     * 'history': things done
    #     * 'all': all tasks
    # person:
    #    * 'all': everyone
    #    * 'string': someone
    # deadline: return only the tasks that will expire by the end of the "period"
    # conditions: 
    #    other conditions, support more than one exprssions, and they are seperated by one spacebar using and logic
    #    the expression is like key1 operator1 value1.1, value1.2
    #    keys are defined in the task DB
    #    operators include =(value equal), <=, >=, >, <, !=, ~=(include), ==(string equal)
    #    there can be more than one value in one expression with or logic, which can be seperate using ', '
    #    e.g. ti==test mo=private jobs st<=2018-09-13
    # detail: how many descriptions will be added
    #    * None: no description
    #    * some: description added during the period
    #    * all: all description
    # sort: sorting order
    #    * module: sort by module
    #    * time:sort by time

    startDate, endDate = GetPeriod(period)

    # assemble the search code using TinyDB grammar
    q = td.Query()
    query = []

    ## period
    if period == 'all': pass
    elif startDate == '': query.append("((q.st == '') | (q.en == ''))")
    else: query.append("(q.st <= endDate) & (q.en >= startDate)")

    ## deadline
    if deadline == True: query.append("(q.en <= endDate)")

    ## range
    if range == 'todo': query.append("(q.ra < 100)")
    elif range == 'history':
        history = logDB.search((q.ti >= startDate) & (q.ti <= endDate) & (q.me == 'u'))
        taskNumberList = []
        for h in history:
            if h['no'] not in taskNumberList: 
                taskNumberList.append(h['no'])            
        
        query.append("(q.no.one_of(taskNumberList))")
    else: pass

    ## person
    if person == 'all': pass
    else: query.append('(q.pe == "' + person + '")')
        
    # assemble the other conditions
    if conditions != None:  query.append( '(' + AssembleQuery(conditions) + ')')

    
    # start to search
    if query == []: taskList = taskDB.all()
    else:
        query = ' & '.join(query)
        taskList = taskDB.search(eval(query))
    
    # handle the relationships between the task, subtask, and parent task
    dataM = []
    numberList = []
    i = 0
    while i < len(taskList):
        
        task = taskList[i]
        i += 1
        
        # the task is already handled
        if task['no'] in numberList: continue

        # prepare for sorting in time
        if sort == 'time':
            if task['en'] > endDate: task['mo'] = '0'
            else: task['mo'] = task['en']

        # handle the detail(description)
        if detail == None:
            task['de'] = ''
        elif detail == 'some':
            task['de'] = ProcessDescription(task, startDate, endDate)
        elif detail == 'all':
            task['de'] = ProcessDescription(task)


        # add all the subtask
        if range == 'all':
            children = taskDB.search(q.pa == task['no'])
            for child in children:
                taskList.insert(i,  child)
            
            if startDate == '' and children != '':
                continue

        # if range != 'todo' and 'pa' in task.keys() and (task['pa'] != '' and task['pa'] != None): # if it's a subtask, only in range "history" and "all"
        if 'pa' in task.keys() and (task['pa'] != '' and task['pa'] != None):

            parent = FindParent(task['pa'], dataM)
            if parent != None: # parent task is already added into the list
                if 'ch' in parent.keys():
                    parent['ch'].append(task)
                else:
                    parent['ch'] = [task]

                numberList.append(task['no'])
            
            else: # add the parent task into the list

                    parent = taskDB.search(q.no == task['pa'])
                    if parent == []:
                        print('C Parent task %s not found'%task['pa'])
                        sys.exit()
                    else:
                        i -= 1
                        taskList.insert(i, parent[0])

        else: # this is not a subtask
            
            dataM.append(task)
            numberList.append(task['no'])

    return dataM
        

def GetTheOrder(task):
    p = ''
    if task['mo'] in smallJobs: p = '2' + '%02d'%smallJobs.index(task['mo'])
    elif task['mo'] in myJobs: p = '3' + '%02d'%myJobs.index(task['mo'])
    elif task['mo'] in importantJobs: p = '0' + '%02d'%importantJobs.index(task['mo'])
    else: p = '100'

    en = task['en'] if task['en'] != '' else '8'
    if task['ra'] == -1: en = '9' # 暂停的项目排在最后

    return p + task['mo'] + en + '%5s'%task['no']


# deal with the json of subtask and convert it to plain text
def SubTaskPlain(list, level, format = 'md list'):

    # sort
    list = sorted(list, key = GetTheOrder)

    # deal with the json
    text = ''
    pList = [] # the list that is handled already, used in excel format
    startDate = None
    endDate = None
    ratio = None
    for i in list:
        
        # deal with he indentation
        j = 0
        tab = ''
        space = ''
        spaceE = ''
        while j < level:
            tab += '  '
            space += '&nbsp;&nbsp;&nbsp;'
            spaceE += '   '
            j += 1

        # deal with the  subtask first
        if 'ch' in i.keys(): 
            t, s, e, r = SubTaskPlain(i['ch'], level + 1, format) 
            # t: the content plained, or a list contained the content plained
            # s: the earliest start time of the subtasks
            # e: the latest end time of the subtasks
            # r: are the subtasks completed
            if e != '' and i['en'] <= e: i['en'] = e
            else: 
                if i['ra'] != -1: i['ra'] = '?' # if the task is not paused, then the ratio depends on its subtasks
            if s != '' and i['st'] >= e: i['st'] = s
            if r != None and i['ra'] != '?' and i['ra'] != -1: i['ra'] = r
        else:
            t = ''

        if 'ra' in i.keys():
            if i['ra'] == -1:
                state = 'Paused'
            elif i['ra'] == '?':
                state = 'Not Scheduled'
            elif i['ra'] == 100:
                state = 'Completed'
            elif i['ra'] == 0:
                state = 'Not Started'
            else:
                state = 'Udergoing'
        else:
            if i['en'] == '':
                state = 'Not Scheduled'
            else:
                state = 'Not Started'
        
        # help to parent task to confirm the start time and end time
        if i['st'] != '' and (startDate == None or startDate > i['st']):
            startDate = i['st']

        if i['en'] == '':
            endDate = ''
        elif endDate != '' and (endDate == None or endDate < i['en']): 
            endDate = i['en']

        
        # deal with the descriptions
        if format == 'md list': # markdown list

            i['de'] = re.sub(r'(=== [0-9]{4}-[0-9]{2}-[0-9]{2} ===\n)', '', i['de'])
            i['de'] = re.sub('---jnv8h line---', '; ', i['de'])
            i['de'] = re.sub(r'(\r\n)+', '; ', i['de'])
            i['de'] = re.sub(r'(\n)+', '; ', i['de'])
            if i['de'] == '': i['de'] = 'RT'

            text += '\n' + tab + '* **' + i['ti'] + '**：' + i['de'] + t
        
        elif format == 'md list no': # markdown list with the task number

            i['de'] = re.sub(r'(=== [0-9]{4}-[0-9]{2}-[0-9]{2} ===\n)', '', i['de'])
            i['de'] = re.sub('---jnv8h line---', '; ', i['de'])
            i['de'] = re.sub(r'(\r\n)+', '；', i['de'])
            i['de'] = re.sub(r'(\n)+', '；', i['de'])
            if i['de'] == '': i['de'] = 'RT'

            text += '\n' + tab + '* **' + i['ti'] + '**：' + i['de'] + ' <!--' + i['no'] + '-->' + t

        elif format == 'md table': # markdown table

            if level != 0: space += '┗'
            de = re.sub(r'---jnv8h line---', '\n', i['de'])
            de = re.sub(r'\n', '<br/>', i['de'])
            text += '\n|' + space + i['ti'] + '|' + de + '|' + i['pe'] + '|' + state + '|' + i['st'] + '|' + i['en'] + '|' + ' <!--' + i['no'] + '-->' + t
        
        elif format == 'excel': # excel

            if i['ra'] != '0' and i['ra'] != '100' and i['ra'] != '-1': ratio = '20'

            if level != 0: spaceE += '┗'
            pList.append({
                'mo': i['mo'],
                'ti': spaceE + i['ti'],
                'de': '="' + i['de'] + '"',
                'pe': i['pe'],
                'ra': state,
                'st': i['st'],
                'en': i['en'],
                'no': i['no'],
                'pr': i['pr'],
                'pa': i['pa']
            })
            pList.extend(t)
      
    if startDate == None: startDate = ''
    if endDate == None: endDate = ''
    if text != '': return text, startDate, endDate, ratio
    else: return pList, startDate, endDate, ratio


# generate the report in md format, list or table
def GenerateMD(taskList, format = 'md list'):
      
    taskList = sorted(taskList, key = GetTheOrder)
    
    result = {}
    for json in taskList:
        
        # make the result plain
        if json['mo'] not in result.keys(): # 模块不存在则添加模块
            result[json['mo']] = ''

        t, s, e, tt = SubTaskPlain([json], 0, format)
        result[json['mo']] += t

    # assemble the md
    text = ''
    for i in result.keys():
        if format == 'md list' or format == 'md list no':
            text += '### ' + i + '\n' + result[i] + '\n\n'
        elif format == 'md table':
            title = '\n|Task|Detail|Person|State|Start time|Planned time|\n|---|---|---|---|---|---|'
            text += '### ' + i + '\n' + title + result[i] + '\n\n'
    
    pyperclip.copy(text)
    return text


def GenerateExcel(taskList, filePath=os.curdir, fileName='report', showParent=False):

    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    # style
    cellString = Side(border_style='thin', color='AAAAAA')
    cellBorder = Border(top=cellString, bottom=cellString, left=cellString, right=cellString)
    cellFill = PatternFill(patternType='solid', fill_type='solid', fgColor='A7C0DE')
    

    # title
    ws['A1'] = 'Module'
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['A'].fill = cellFill
    ws.column_dimensions['A'].border = cellBorder
    ws['A1'].fill = cellFill
    ws['A1'].border = cellBorder
    ws['B1'] = 'Task'
    ws.column_dimensions['B'].width = 36
    ws['B1'].fill = cellFill
    ws['B1'].border = cellBorder
    ws['C1'] = 'Person'
    ws.column_dimensions['C'].width = 12
    ws['C1'].fill = cellFill
    ws['C1'].border = cellBorder
    ws['D1'] = 'State'
    ws.column_dimensions['D'].width = 12
    ws['D1'].fill = cellFill
    ws['D1'].border = cellBorder
    ws['E1'] = 'Start Time'
    ws.column_dimensions['E'].width = 18
    ws['E1'].fill = cellFill
    ws['E1'].border = cellBorder
    ws['F1'] = 'Planned Time'
    ws.column_dimensions['F'].width = 18
    ws['F1'].fill = cellFill
    ws['F1'].border = cellBorder
    ws['G1'] = 'Detail'
    ws.column_dimensions['G'].width = 36
    ws['G1'].fill = cellFill
    ws['G1'].border = cellBorder
    ws['H1'] = 'Number'
    ws['H1'].fill = cellFill
    ws['H1'].border = cellBorder
    ws['I1'] = 'Priority'
    ws['I1'].fill = cellFill
    ws['I1'].border = cellBorder
    if showParent == True:
        ws['J1'] = 'Parent No'
        ws['J1'].fill = cellFill
        ws['J1'].border = cellBorder


    rowList, s, e, tt = SubTaskPlain(taskList, level = 0, format = 'excel')
    
    # start to write the data
    i = 1 # current line
    j = 0 # the cells to be merged
    mo = ''
    for row in rowList:
        i += 1
        rowNumber = str(i)

        if row['mo'] == '': row['mo'] = 'Default'
        j += 1
        if row['mo'] != mo: # new module that haven't appeared
            mo = row['mo']
            if j != 1: # merge the cells 
                ws.merge_cells(start_row = i - j, start_column = 1, end_row = i - 1, end_column = 1)
            ws['A' + rowNumber] = row['mo']
            ws['A' + rowNumber].alignment = Alignment(vertical = 'top', wrap_text = True)
            ws['A' + rowNumber].fill = cellFill
            ws['A' + rowNumber].border = cellBorder
            j = 0

        if i == len(rowList) + 1: # list bottom
            ws.merge_cells(start_row = i - j, start_column = 1, end_row = i, end_column = 1)
            pass

        ws['B' + rowNumber] = row['ti']
        ws['B' + rowNumber].alignment = Alignment(vertical = 'top', wrap_text = True)
        ws['C' + rowNumber] = row['pe']
        ws['C' + rowNumber].alignment = Alignment(vertical = 'top')
        ws['D' + rowNumber] = row['ra']
        ws['D' + rowNumber].alignment = Alignment(vertical = 'top')
        ws['E' + rowNumber] = row['st']
        ws['E' + rowNumber].alignment = Alignment(vertical = 'top')
        ws['F' + rowNumber] = row['en']
        ws['F' + rowNumber].alignment = Alignment(vertical = 'top')
        ws['G' + rowNumber] = row['de']
        ws['G' + rowNumber].alignment = Alignment(vertical = 'top')
        ws['G' + rowNumber].alignment = Alignment(wrap_text = True)
        ws['H' + rowNumber] = row['no']
        ws['H' + rowNumber].alignment = Alignment(vertical = 'top')
        ws['I' + rowNumber] = row['pr']
        ws['I' + rowNumber].alignment = Alignment(vertical = 'top')
        if showParent == True: 
            ws['J' + rowNumber] = row['pa']
            ws['J' + rowNumber].alignment = Alignment(vertical = 'top')
        
        # format
        if not re.search('┗', row['ti']): 
            ws['B' + rowNumber].font = Font(bold = True)#, color = 'FF0000')
            ws['C' + rowNumber].font = Font(bold = True)#, color = 'FF0000')
            if row['ra'] == 'Not Scheduled':
                ws['D' + rowNumber].font = Font(bold = True, color = 'FF0000')
            else:
                ws['D' + rowNumber].font = Font(bold = True)#, color = 'FF0000')
            ws['E' + rowNumber].font = Font(bold = True)#, color = 'FF0000')
            ws['F' + rowNumber].font = Font(bold = True)#, color = 'FF0000')
            # ws['G' + rowNumber].font = Font(bold = True)#, color = 'FF0000')
            ws['H' + rowNumber].font = Font(bold = True)#, color = 'FF0000')
            ws['I' + rowNumber].font = Font(bold = True)#, color = 'FF0000')
        else:
            ws['B' + rowNumber].font = Font(color = '999999')
            ws['C' + rowNumber].font = Font(color = '999999')
            ws['D' + rowNumber].font = Font(color = '999999')
            ws['E' + rowNumber].font = Font(color = '999999')
            ws['F' + rowNumber].font = Font(color = '999999')
            ws['G' + rowNumber].font = Font(color = '999999')
            ws['H' + rowNumber].font = Font(color = '999999')
            ws['I' + rowNumber].font = Font(color = '999999')

    wb.save( filePath + os.sep + fileName + '.xlsx')
    return 'Saved in ' + filePath + os.sep + fileName + '.xlsx'


def AssembleQuery(args):

    operators = r'=|<=|>=|>|<|!=|~=|=='
    reg = r'\s*([a-zA-Z0-9]+)(' + operators + ')'
    kwargs = []

    match = re.search(reg, args)
    ve = vs = 0
    while match != None:
        matchf = match
        k = matchf.group(1) # keyword
        o = matchf.group(2) # operator
        vs = matchf.span()[1] + vs - ve
        match = re.search(reg, args[vs:])
        ve = None if match == None else match.span()[0] + vs
        v = args[vs: ve] # value
        args = args[ve:]
        kwargs.append([k, o, v])

    conditions = []
    for [k, o, v] in kwargs:
        vList = v.split(', ') # supprt the or logical
        ct = []
        for v in vList:
            if v == '': continue

            if o == '~=':
                ct.append("q." + k + ".search('" + v + "')")
            elif o == '!=':
                ct.append("~q." + k + ".search('" + v + "')")
            elif o == '=':
                ct.append("q." + k + " == '" + v + "'")
            elif o == '==':
                ct.append("q." + k + " == " + v)
            else:
                ct.append("q." + k + o + v)
        
        if len(ct) == 1:
            c = ct[0]
        else:
            c = '(' + ') | ('.join(ct) + ')'

        conditions.append(c)
    
    if len(conditions) == 0:
        return ''
    elif len(conditions) == 1:
        c = conditions[0]
    else:
        c = '(' + ') & ('.join(conditions) + ')'
    
    return c


def Search(args):

    # You can add any other conditions if you want to filt the result.
    # The expression includs three parts, e.g. "ti==test"
    #
    #    * keys: keys defined in the task DB, "ti" in the example
    #    * operator:  includes =(value equal), <=, >=, >, <, !=, ~=(include), ==(string equal), "==" in the example
    #    * value: "test" in the example
    # 
    # There can be more than one value in one expression with "or" logic, which can be seperate using ', ', e.g. "ti==test1, test2"
    # NOTE: spacebar is allowed between the parts, that is "ti == test1, test2" is not illegal, but "ti==test of one, test of two" is legal
    #
    # And it support more than one exprssions in one sentence, these expressions should be seperated by one spacebar, and using "and" logic
    # e.g. "ti==test mo=private jobs st<=2018-09-13"

    c = AssembleQuery(args)
    q = td.Query()
    if c != '': taskList = taskDB.search(eval(c))
    else: taskList =  taskDB.all()
    
    # filter the result
    j = 0 # 用于抵消被删除的元素带来的长度变化影响
    for i in range(len(taskList)):
        task = taskList[i - j]

        # 编辑优先级
        if 'pr' not in task.keys() or task['pr'] == None: 
            task['pr'] = '3'
    
    taskList = sorted(taskList, key = lambda a: a['pr'])

    return taskList